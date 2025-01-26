import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']
driver = webdriver.Chrome()
driver.get('site')

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha 
    sleep(5)
#Sempre verificar no navegador se o id da tag que está sendo utilizado é único
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='value']")
    sleep(1)
    campo_pesquisa.clear
    campo_pesquisa.send_keys(cpf)
    sleep(1)
#Verificar se está "em dia" ou "atrasado"
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
#Se estiver "em dia", pegar a data do pagamento e o método
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    status.text
    if status.text == "em dia":
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.split()[3]

        planilha_fechamento = openpyxl.load_workbook('nome_planilha')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf, vencimento,'em dia',data_pagamento_limpo,metodo_pagamento_limpo])
        planilha_fechamento.save('nome_planilha')
    else:
        planilha_fechamento = openpyxl.load_workbook('nome_planilha')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente'])
